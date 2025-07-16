import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import itertools

PERIODOS_POR_DIA = 7
ARCHIVO_EXCEL = "./horario_completo_generado.xlsx"
ARCHIVO_MAESTROS = "./1.json"
ARCHIVO_PREFERENCIAS = "./preferencias.json"
ARCHIVO_SALIDA = "horario_ajustado.xlsx"

DIAS_MAP = {
    "Lunes": "LUNES",
    "lunes": "LUNES",
    "Martes": "MARTES",
    "martes": "MARTES",
    "Miercoles": "MIÉRCOLES",
    "miercoles": "MIÉRCOLES",
    "Miércoles": "MIÉRCOLES",
    "Jueves": "JUEVES",
    "jueves": "JUEVES",
    "Viernes": "VIERNES",
    "viernes": "VIERNES",
}
DIAS_ORDENADOS = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]

HORAS_POR_ANO = {
    "1": {
        "español": 5,
        "ingles": 3,
        "matematicas": 5,
        "ciencias": 4,
        "historia": 2,
        "geografia": 4,
        "formacion civica y etica": 2,
        "educacion artistica": 3,
        "tutoria": 1,
        "educacion fisica": 2,
        "tecnologia": 3,
        "integracion": 1,
    },
    "2": {
        "español": 5,
        "ingles": 3,
        "matematicas": 5,
        "ciencias": 6,
        "historia": 4,
        "formacion civica y etica": 2,
        "educacion artistica": 3,
        "tutoria": 1,
        "educacion fisica": 2,
        "tecnologia": 3,
        "integracion": 1,
    },
    "3": {
        "español": 5,
        "ingles": 3,
        "matematicas": 5,
        "ciencias": 6,
        "historia": 4,
        "formacion civica y etica": 2,
        "tutoria": 1,
        "educacion artistica": 3,
        "educacion fisica": 2,
        "tecnologia": 3,
        "integracion": 1,
    },
}


class AjustadorHorarios:
    def __init__(self):
        self.horarios_grupos = {}
        self.horarios_maestros = {}
        self.maestros_data = None
        self.preferencias = {}
        self.horas_ordenadas = None
        # NUEVO: Mapeo entre nombres de hojas y nombres cortos para consistencia
        self.maestro_sheet_to_short = {}
        self.short_to_maestro_sheet = {}
        self._cargar_datos()

    def _crear_mapeo_nombres(self):
        """Crear mapeo consistente entre nombres de hojas y nombres cortos"""
        for maestro_data in self.maestros_data:
            nombre_completo = maestro_data["nombre"]
            maestro_sheet_key = self._encontrar_maestro_en_horarios(nombre_completo)
            if maestro_sheet_key:
                short_name = self._get_short_name(nombre_completo)
                self.maestro_sheet_to_short[maestro_sheet_key] = short_name
                self.short_to_maestro_sheet[short_name] = maestro_sheet_key

    def _esta_maestro_ocupado(self, maestro_referencia, dia, hora):
        """Verifica si un maestro ya está dando clase en otro grupo en un momento específico."""
        # CORRECCIÓN: Normalizar la referencia del maestro
        if maestro_referencia in self.short_to_maestro_sheet:
            maestro_sheet_key = self.short_to_maestro_sheet[maestro_referencia]
        elif maestro_referencia in self.maestro_sheet_to_short:
            maestro_sheet_key = maestro_referencia
        else:
            # Si no se encuentra en el mapeo, usar el nombre tal como viene
            maestro_sheet_key = maestro_referencia

        for grupo, df_grupo in self.horarios_grupos.items():
            if dia in df_grupo.columns and hora in df_grupo.index:
                celda = df_grupo.at[hora, dia]
                if pd.notna(celda) and isinstance(celda, str):
                    partes = celda.split("\n")
                    if len(partes) > 1:
                        maestro_en_celda = partes[1].strip()
                        # Comparar usando el mapeo consistente
                        if maestro_en_celda == maestro_sheet_key:
                            return True
        return False

    def _viola_limite_consecutivo(self, grupo, dia, hora, materia, limite=2):
        """Verifica si añadir una clase crearía más de `limite` horas consecutivas."""
        df_grupo = self.horarios_grupos[grupo]
        materia_norm = materia.upper()

        df_grupo_simulado = df_grupo.copy()
        df_grupo_simulado.at[hora, dia] = f"{materia_norm}\nMAESTRO_SIMULADO"

        consecutivas = 0
        for h in self.horas_ordenadas:
            celda = df_grupo_simulado.at[h, dia]
            materia_en_celda = (
                str(celda).split("\n")[0].upper() if pd.notna(celda) else ""
            )
            if materia_en_celda == materia_norm:
                consecutivas += 1
                if consecutivas > limite:
                    return True
            else:
                consecutivas = 0
        return False

    def _crea_clase_dividida(self, grupo, dia, hora, materia):
        """Evita que una materia quede con un hueco en medio. Ej: MAT | HUECO | MAT"""
        df_grupo = self.horarios_grupos[grupo]
        materia_norm = materia.upper()
        idx = self.horas_ordenadas.index(hora)

        if idx > 1:
            hora_anterior = self.horas_ordenadas[idx - 1]
            hora_ante_anterior = self.horas_ordenadas[idx - 2]
            celda_anterior = df_grupo.at[hora_anterior, dia]
            celda_ante_anterior = df_grupo.at[hora_ante_anterior, dia]
            materia_ante_anterior = (
                str(celda_ante_anterior).split("\n")[0].upper()
                if pd.notna(celda_ante_anterior)
                else ""
            )
            if pd.isna(celda_anterior) and materia_ante_anterior == materia_norm:
                return True

        if idx < len(self.horas_ordenadas) - 2:
            hora_siguiente = self.horas_ordenadas[idx + 1]
            hora_post_siguiente = self.horas_ordenadas[idx + 2]
            celda_siguiente = df_grupo.at[hora_siguiente, dia]
            celda_post_siguiente = df_grupo.at[hora_post_siguiente, dia]
            materia_post_siguiente = (
                str(celda_post_siguiente).split("\n")[0].upper()
                if pd.notna(celda_post_siguiente)
                else ""
            )
            if pd.isna(celda_siguiente) and materia_post_siguiente == materia_norm:
                return True

        return False

    def _viola_limite_bloques_dobles(self, grupo, dia, hora, materia, limite=1):
        """Verifica que una materia no tenga más de `limite` bloques de 2 horas en un día."""
        df_grupo = self.horarios_grupos[grupo]
        materia_norm = materia.upper()

        df_grupo_simulado = df_grupo.copy()
        df_grupo_simulado.at[hora, dia] = f"{materia_norm}\nMAESTRO_SIMULADO"

        bloques_dobles = 0
        i = 0
        while i < len(self.horas_ordenadas) - 1:
            materia1 = (
                str(df_grupo_simulado.at[self.horas_ordenadas[i], dia])
                .split("\n")[0]
                .upper()
                if pd.notna(df_grupo_simulado.at[self.horas_ordenadas[i], dia])
                else ""
            )
            materia2 = (
                str(df_grupo_simulado.at[self.horas_ordenadas[i + 1], dia])
                .split("\n")[0]
                .upper()
                if pd.notna(df_grupo_simulado.at[self.horas_ordenadas[i + 1], dia])
                else ""
            )
            if materia1 == materia_norm and materia2 == materia_norm:
                bloques_dobles += 1
                i += 2
            else:
                i += 1
        return bloques_dobles > limite

    def _cargar_datos(self):
        with open(ARCHIVO_MAESTROS, "r", encoding="utf-8") as f:
            self.maestros_data = json.load(f)
        with open(ARCHIVO_PREFERENCIAS, "r", encoding="utf-8") as f:
            prefs_list = json.load(f)
            for pref in prefs_list:
                nombre = pref["nombre"]
                horarios = {
                    DIAS_MAP.get(dia, dia.upper()): [
                        p - 1 for p in pref["horarios"][dia]
                    ]
                    for dia in pref["horarios"]
                }
                self.preferencias[nombre] = horarios

        xls = pd.ExcelFile(ARCHIVO_EXCEL)
        for sheet in xls.sheet_names:
            if not (sheet.startswith("Grupo ") or sheet.startswith("Maestro ")):
                continue

            df_raw = pd.read_excel(xls, sheet_name=sheet, index_col=0)
            if df_raw.empty:
                continue

            if self.horas_ordenadas is None and not df_raw.empty:
                self.horas_ordenadas = df_raw.index.tolist()[:PERIODOS_POR_DIA]

            if self.horas_ordenadas is None:
                continue

            df_clean = pd.DataFrame(index=self.horas_ordenadas, columns=DIAS_ORDENADOS)
            col_rename_map = {
                DIAS_MAP.get(col, str(col).upper()): col for col in df_raw.columns
            }

            for periodo_actual in self.horas_ordenadas:
                if periodo_actual in df_raw.index:
                    row = df_raw.loc[periodo_actual]
                    for dia_norm in DIAS_ORDENADOS:
                        col_orig = col_rename_map.get(dia_norm)
                        if col_orig and col_orig in row and pd.notna(row[col_orig]):
                            df_clean.at[periodo_actual, dia_norm] = str(
                                row[col_orig]
                            ).strip()

            if sheet.startswith("Grupo "):
                self.horarios_grupos[sheet.replace("Grupo ", "")] = df_clean
            elif sheet.startswith("Maestro "):
                self.horarios_maestros[sheet.replace("Maestro ", "")] = df_clean

        # NUEVO: Crear mapeo después de cargar todos los datos
        self._crear_mapeo_nombres()

    def _get_short_name(self, full_name):
        return " ".join(full_name.strip().split()[:2])

    def _encontrar_maestro_en_horarios(self, full_name_pref):
        norm_full_name = full_name_pref.upper().strip()
        possible_matches = []
        for key_sheet in self.horarios_maestros.keys():
            norm_key_sheet = key_sheet.upper().strip()
            if norm_key_sheet == norm_full_name:
                return key_sheet
            if norm_key_sheet in norm_full_name:
                possible_matches.append(key_sheet)
        if possible_matches:
            return max(possible_matches, key=len)
        return None

    def _encontrar_slot_cercano(
        self, dia, periodo_pref, grupo, materia, maestro_sheet_key, df_grupo, df_maestro
    ):
        posibles = []
        for offset in range(1, PERIODOS_POR_DIA):
            for sign in [-1, 1]:
                p = periodo_pref + sign * offset
                if 0 <= p < PERIODOS_POR_DIA:
                    hora_p = self.horas_ordenadas[p]
                    if (
                        pd.isna(df_grupo.at[hora_p, dia])
                        and pd.isna(df_maestro.at[hora_p, dia])
                        and not self._esta_maestro_ocupado(
                            maestro_sheet_key, dia, hora_p
                        )
                        and not self._viola_limite_consecutivo(
                            grupo, dia, hora_p, materia
                        )
                        and not self._crea_clase_dividida(grupo, dia, hora_p, materia)
                        and not self._viola_limite_bloques_dobles(
                            grupo, dia, hora_p, materia
                        )
                    ):
                        posibles.append((abs(offset), p))
        if posibles:
            return sorted(posibles)[0][1]
        return None

    def ajustar_preferencias(self):
        for full_name, prefs in self.preferencias.items():
            maestro_sheet_key = self._encontrar_maestro_en_horarios(full_name)
            if not maestro_sheet_key:
                continue

            df_maestro = self.horarios_maestros[maestro_sheet_key]

            clases_a_mover = []
            for dia in DIAS_ORDENADOS:
                for p in range(PERIODOS_POR_DIA):
                    hora_p = self.horas_ordenadas[p]
                    cell = df_maestro.at[hora_p, dia]
                    if pd.notna(cell) and str(cell).strip() not in ["", "SERVICIO"]:
                        grupo_en_celda = str(cell).strip()
                        grupo = grupo_en_celda.split("\n")[-1].strip("()")

                        if grupo in self.horarios_grupos:
                            celda_grupo = self.horarios_grupos[grupo].at[hora_p, dia]
                            if pd.notna(celda_grupo):
                                materia = str(celda_grupo).split("\n")[0]
                                clases_a_mover.append((dia, p, grupo, materia))

            for dia_orig, p_orig, grupo, materia in clases_a_mover:
                if materia.upper() == "TECNOLOGIA":
                    continue
                prefs_dia = prefs.get(dia_orig, [])
                if p_orig not in prefs_dia:
                    slot_encontrado = False
                    prefs_dia_sorted = sorted(prefs_dia, key=lambda x: abs(x - p_orig))
                    for pref_p in prefs_dia_sorted:
                        hora_pref = self.horas_ordenadas[pref_p]
                        if (
                            pd.isna(self.horarios_grupos[grupo].at[hora_pref, dia_orig])
                            and pd.isna(df_maestro.at[hora_pref, dia_orig])
                            and not self._esta_maestro_ocupado(
                                maestro_sheet_key, dia_orig, hora_pref
                            )
                            and not self._viola_limite_consecutivo(
                                grupo, dia_orig, hora_pref, materia
                            )
                            and not self._crea_clase_dividida(
                                grupo, dia_orig, hora_pref, materia
                            )
                            and not self._viola_limite_bloques_dobles(
                                grupo, dia_orig, hora_pref, materia
                            )
                        ):
                            hora_orig = self.horas_ordenadas[p_orig]
                            self.horarios_grupos[grupo].at[hora_orig, dia_orig] = None
                            df_maestro.at[hora_orig, dia_orig] = None
                            # CORRECCIÓN: Usar maestro_sheet_key consistentemente
                            self.horarios_grupos[grupo].at[hora_pref, dia_orig] = (
                                f"{materia}\n{maestro_sheet_key}"
                            )
                            df_maestro.at[hora_pref, dia_orig] = grupo
                            slot_encontrado = True
                            break

                    if slot_encontrado:
                        continue

                    for pref_p in prefs_dia_sorted:
                        cercano_p = self._encontrar_slot_cercano(
                            dia_orig,
                            pref_p,
                            grupo,
                            materia,
                            maestro_sheet_key,
                            self.horarios_grupos[grupo],
                            df_maestro,
                        )
                        if cercano_p is not None:
                            hora_orig = self.horas_ordenadas[p_orig]
                            hora_cercana = self.horas_ordenadas[cercano_p]
                            self.horarios_grupos[grupo].at[hora_orig, dia_orig] = None
                            df_maestro.at[hora_orig, dia_orig] = None
                            # CORRECCIÓN: Usar maestro_sheet_key consistentemente
                            self.horarios_grupos[grupo].at[hora_cercana, dia_orig] = (
                                f"{materia}\n{maestro_sheet_key}"
                            )
                            df_maestro.at[hora_cercana, dia_orig] = grupo
                            break

    def _backtrack_compactar(
        self, clases, dia, maestro_sheet_key, posiciones_posibles, index=0
    ):
        if index == len(clases):
            return True

        grupo, materia, _ = clases[index]
        df_grupo = self.horarios_grupos[grupo]
        df_maestro = self.horarios_maestros[maestro_sheet_key]

        for p in posiciones_posibles:
            hora_p = self.horas_ordenadas[p]
            if (
                pd.isna(df_grupo.at[hora_p, dia])
                and pd.isna(df_maestro.at[hora_p, dia])
                and not self._esta_maestro_ocupado(maestro_sheet_key, dia, hora_p)
                and not self._viola_limite_consecutivo(grupo, dia, hora_p, materia)
                and not self._crea_clase_dividida(grupo, dia, hora_p, materia)
                and not self._viola_limite_bloques_dobles(grupo, dia, hora_p, materia)
            ):
                df_grupo.at[hora_p, dia] = f"{materia}\n{maestro_sheet_key}"
                df_maestro.at[hora_p, dia] = grupo

                if self._backtrack_compactar(
                    clases,
                    dia,
                    maestro_sheet_key,
                    posiciones_posibles,
                    index + 1,
                ):
                    return True

                df_grupo.at[hora_p, dia] = None
                df_maestro.at[hora_p, dia] = None

        return False

    def compactar_horarios_maestros(self):
        if self.horas_ordenadas is None:
            return

        for maestro_data in self.maestros_data:
            maestro_sheet_key = self._encontrar_maestro_en_horarios(
                maestro_data["nombre"]
            )
            if not maestro_sheet_key:
                continue

            df_maestro = self.horarios_maestros[maestro_sheet_key]

            for dia in DIAS_ORDENADOS:
                clases_a_compactar = []

                for p_idx, hora_p in enumerate(self.horas_ordenadas):
                    cell = df_maestro.at[hora_p, dia]
                    if (
                        pd.notna(cell)
                        and str(cell).strip() not in ["", "SERVICIO"]
                        and "(" in str(cell)
                    ):
                        grupo = str(cell).strip().split("\n")[-1].strip("()")

                        materia_celda = self.horarios_grupos[grupo].at[hora_p, dia]
                        if pd.notna(materia_celda):
                            materia = materia_celda.split("\n")[0]
                            clase_info = (grupo, materia, p_idx)

                            if materia.upper() == "TECNOLOGIA":
                                pass
                            else:
                                clases_a_compactar.append(clase_info)
                                self.horarios_grupos[grupo].at[hora_p, dia] = None
                                df_maestro.at[hora_p, dia] = None

                if not clases_a_compactar:
                    continue

                num_clases_movibles = len(clases_a_compactar)

                posibles_bloques = []
                for start in range(PERIODOS_POR_DIA - num_clases_movibles + 1):
                    bloque = list(range(start, start + num_clases_movibles))
                    if all(
                        pd.isna(df_maestro.at[self.horas_ordenadas[p], dia])
                        for p in bloque
                    ):
                        posibles_bloques.append(bloque)

                prefs_dia = self.preferencias.get(maestro_data["nombre"], {}).get(
                    dia, []
                )
                posibles_bloques.sort(
                    key=lambda b: -sum(1 for p in b if p in prefs_dia)
                )

                compactado = False
                for bloque in posibles_bloques:
                    for perm in itertools.permutations(clases_a_compactar):
                        valid = True
                        for i, (grupo, materia, _) in enumerate(perm):
                            p_idx = bloque[i]
                            hora_p = self.horas_ordenadas[p_idx]

                            if (
                                self._esta_maestro_ocupado(
                                    maestro_sheet_key, dia, hora_p
                                )
                                or self._viola_limite_consecutivo(
                                    grupo, dia, hora_p, materia
                                )
                                or self._crea_clase_dividida(
                                    grupo, dia, hora_p, materia
                                )
                                or self._viola_limite_bloques_dobles(
                                    grupo, dia, hora_p, materia
                                )
                            ):
                                valid = False
                                break

                        if valid:
                            for i, (grupo, materia, _) in enumerate(perm):
                                p_idx = bloque[i]
                                hora_p = self.horas_ordenadas[p_idx]
                                # CORRECCIÓN: Usar maestro_sheet_key consistentemente
                                self.horarios_grupos[grupo].at[hora_p, dia] = (
                                    f"{materia}\n{maestro_sheet_key}"
                                )
                                df_maestro.at[hora_p, dia] = grupo
                            compactado = True
                            break
                    if compactado:
                        break

                if not compactado:
                    for grupo, materia, p_orig_idx in clases_a_compactar:
                        hora_orig = self.horas_ordenadas[p_orig_idx]
                        # CORRECCIÓN: Usar maestro_sheet_key consistentemente
                        self.horarios_grupos[grupo].at[hora_orig, dia] = (
                            f"{materia}\n{maestro_sheet_key}"
                        )
                        df_maestro.at[hora_orig, dia] = grupo

    def llenar_huecos_servicio(self):
        for maestro_data in self.maestros_data:
            maestro_sheet_key = self._encontrar_maestro_en_horarios(
                maestro_data["nombre"]
            )
            if not maestro_sheet_key:
                continue

            df_maestro = self.horarios_maestros[maestro_sheet_key]
            horas_servicio_disp = maestro_data.get("horas_servicio", 0) or 0

            for dia in DIAS_ORDENADOS:
                clases_indices = [
                    p
                    for p in range(PERIODOS_POR_DIA)
                    if pd.notna(df_maestro.at[self.horas_ordenadas[p], dia])
                    and str(df_maestro.at[self.horas_ordenadas[p], dia]).strip()
                    not in ["", "SERVICIO"]
                ]
                if not clases_indices:
                    continue
                primera_clase_idx, ultima_clase_idx = (
                    min(clases_indices),
                    max(clases_indices),
                )

                for p in range(primera_clase_idx + 1, ultima_clase_idx):
                    if horas_servicio_disp <= 0:
                        break
                    hora_p = self.horas_ordenadas[p]
                    if (
                        pd.isna(df_maestro.at[hora_p, dia])
                        or df_maestro.at[hora_p, dia].strip() == ""
                    ):
                        df_maestro.at[hora_p, dia] = "SERVICIO"
                        horas_servicio_disp -= 1

    def guardar(self):
        with pd.ExcelWriter(ARCHIVO_SALIDA, engine="openpyxl") as writer:
            for grupo, df in self.horarios_grupos.items():
                df.to_excel(writer, sheet_name=f"Grupo {grupo}")
            for maestro, df in self.horarios_maestros.items():
                df.to_excel(writer, sheet_name=f"Maestro {maestro}")

        wb = load_workbook(ARCHIVO_SALIDA)

        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        bold_font = Font(bold=True)

        for grupo, df in self.horarios_grupos.items():
            sheet_name = f"Grupo {grupo}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ano = grupo[0]
                if ano not in HORAS_POR_ANO:
                    continue
                horas_requeridas = HORAS_POR_ANO[ano]
                horas_asignadas = {m: 0 for m in horas_requeridas}
                for _, row in df.iterrows():
                    for cell_value in row:
                        if pd.notna(cell_value):
                            materia = str(cell_value).split("\n")[0].strip().lower()
                            if materia in horas_asignadas:
                                horas_asignadas[materia] += 1

                stats_data = {
                    "Materia": list(horas_requeridas.keys()),
                    "Horas Requeridas": list(horas_requeridas.values()),
                    "Horas Asignadas": [
                        horas_asignadas.get(m, 0) for m in horas_requeridas
                    ],
                    "Faltante/Sobrante": [
                        horas_asignadas.get(m, 0) - horas_requeridas[m]
                        for m in horas_requeridas
                    ],
                }
                stats_df = pd.DataFrame(stats_data)
                start_row = ws.max_row + 3

                ws.cell(
                    row=start_row - 1,
                    column=1,
                    value=f"Resumen de Horas - Grupo {grupo}",
                ).font = bold_font

                for r in dataframe_to_rows(stats_df, index=False, header=True):
                    ws.append(r)

                for cell in ws[start_row]:
                    cell.font = bold_font

        for maestro_data in self.maestros_data:
            maestro_key = self._encontrar_maestro_en_horarios(maestro_data["nombre"])
            if not maestro_key:
                continue

            sheet_name = f"Maestro {maestro_key}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                df_maestro = self.horarios_maestros[maestro_key]

                horas_totales = maestro_data.get("horas_grupo", 0) or 0
                horas_servicio = maestro_data.get("horas_servicio", 0) or 0
                horas_tutoria = maestro_data.get("horas_tutoria", 0) or 0

                horas_asignadas_clase = 0
                horas_asignadas_servicio = 0
                horas_asignadas_tutoria = 0
                for _, row in df_maestro.iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            val = str(cell).strip()
                            if "(" in val:
                                horas_asignadas_clase += 1
                            elif val == "SERVICIO":
                                horas_asignadas_servicio += 1
                            elif val == "TUTORIA":
                                horas_asignadas_tutoria += 1

                stats_data = {
                    "Concepto": [
                        "Horas de Clase",
                        "Horas de Tutoria",
                        "Horas de Servicio",
                    ],
                    "Total Requerido": [horas_totales, horas_tutoria, horas_servicio],
                    "Total Asignado": [
                        horas_asignadas_clase,
                        horas_asignadas_tutoria,
                        horas_asignadas_servicio,
                    ],
                    "Faltante/Sobrante": [
                        horas_asignadas_clase - horas_totales,
                        horas_asignadas_tutoria - horas_tutoria,
                        horas_asignadas_servicio - horas_servicio,
                    ],
                }
                stats_df = pd.DataFrame(stats_data)

                start_row = ws.max_row + 3
                ws.cell(
                    row=start_row - 1,
                    column=1,
                    value=f"Resumen de Horas - {maestro_key}",
                ).font = bold_font
                for r in dataframe_to_rows(stats_df, index=False, header=True):
                    ws.append(r)
                for cell in ws[start_row]:
                    cell.font = bold_font

        maestro_key_to_prefs = {
            key: self.preferencias[name]
            for name, prefs in self.preferencias.items()
            if (key := self._encontrar_maestro_en_horarios(name))
        }
        col_map = {day: chr(ord("B") + i) for i, day in enumerate(DIAS_ORDENADOS)}

        for ws in wb.worksheets:
            if ws.title.startswith("Maestro "):
                maestro_key = ws.title.replace("Maestro ", "")
                if maestro_key in maestro_key_to_prefs:
                    prefs = maestro_key_to_prefs[maestro_key]
                    for dia_pref, periodos_pref in prefs.items():
                        if dia_pref in col_map:
                            col_letter = col_map[dia_pref]
                            for p_pref in periodos_pref:
                                if p_pref < PERIODOS_POR_DIA:
                                    cell = ws[f"{col_letter}{p_pref + 2}"]
                                    if cell.value:
                                        cell.fill = green_fill

            elif ws.title.startswith("Grupo "):
                for row in ws.iter_rows(
                    min_row=2,
                    max_row=PERIODOS_POR_DIA + 1,
                    min_col=2,
                    max_col=len(DIAS_ORDENADOS) + 1,
                ):
                    for cell in row:
                        if (
                            cell.value
                            and isinstance(cell.value, str)
                            and "\n" in cell.value
                        ):
                            maestro_key = cell.value.split("\n")[1].strip()
                            if maestro_key in maestro_key_to_prefs:
                                prefs = maestro_key_to_prefs[maestro_key]
                                dia_actual = ws.cell(row=1, column=cell.column).value
                                periodo_actual = cell.row - 2
                                if (
                                    dia_actual in prefs
                                    and periodo_actual in prefs[dia_actual]
                                ):
                                    cell.fill = green_fill

        wb.save(ARCHIVO_SALIDA)
        print(
            f"Archivo ajustado, con estadísticas y coloreado guardado en {ARCHIVO_SALIDA}"
        )

    def run(self):
        print("Iniciando ajuste de horarios...")
        if not self.horarios_grupos and not self.horarios_maestros:
            print("ERROR: No se cargaron datos. Revisa el formato del archivo Excel.")
            return

        self.ajustar_preferencias()
        print("Preferencias de maestros ajustadas.")
        self.compactar_horarios_maestros()
        print("Horarios de maestros compactados (backtracking aplicado).")
        self.llenar_huecos_servicio()
        print("Huecos de servicio llenados.")
        self.guardar()


if __name__ == "__main__":
    ajustador = AjustadorHorarios()
    ajustador.run()
