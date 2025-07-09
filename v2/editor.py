import openpyxl
import pandas as pd
import os
import re
import copy

DIAS = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES"]
COLUMNAS_HORARIO = ["HORA"] + DIAS
VALOR_VACIO = "---"
PERIODOS_POR_DIA = 7


def limpiar_pantalla():
    os.system("cls" if os.name == "nt" else "clear")


def sanitizar_nombre_hoja(nombre):
    nombre = re.sub(r'[\\/*?:"<>|]', "", nombre)
    return nombre[:31]


class EditorHorarios:
    def __init__(self, archivo_excel):
        self.archivo_excel = archivo_excel
        self.horarios_grupos = {}
        self.horarios_maestros = {}
        self.backup_horarios = None
        self.df_analisis = None
        try:
            with open(archivo_excel, "r"):
                pass
        except FileNotFoundError:
            print(f"Error: No se encontró el archivo '{archivo_excel}'.")
            exit()
        self._cargar_datos_manualmente()

    def _cargar_datos_manualmente(self):
        try:
            wb = openpyxl.load_workbook(self.archivo_excel, data_only=True)
        except Exception as e:
            print(f"Error al abrir el archivo Excel con openpyxl: {e}")
            exit()
        for nombre_hoja in wb.sheetnames:
            if nombre_hoja == "Analisis Horas Servicio":
                xls = pd.ExcelFile(self.archivo_excel)
                self.df_analisis = xls.parse(nombre_hoja)
                continue
            ws = wb[nombre_hoja]
            headers_map = {
                str(cell.value).strip().upper(): col_idx + 1
                for col_idx, cell in enumerate(ws[1])
                if cell.value is not None
            }
            data_template = {
                col: [VALOR_VACIO] * PERIODOS_POR_DIA for col in COLUMNAS_HORARIO
            }
            data_template["HORA"] = [
                ws.cell(row=i, column=headers_map.get("HORA", 1)).value
                for i in range(2, 2 + PERIODOS_POR_DIA)
            ]
            df = pd.DataFrame(data_template)
            for row_idx_df, row_num_excel in enumerate(range(2, 2 + PERIODOS_POR_DIA)):
                for col_name_std in DIAS:
                    if col_name_std in headers_map:
                        col_idx_excel = headers_map[col_name_std]
                        cell_value = ws.cell(
                            row=row_num_excel, column=col_idx_excel
                        ).value
                        df.at[row_idx_df, col_name_std] = (
                            cell_value if cell_value is not None else VALOR_VACIO
                        )
            if nombre_hoja.startswith("Grupo "):
                self.horarios_grupos[nombre_hoja.replace("Grupo ", "")] = df
            else:
                self.horarios_maestros[nombre_hoja] = df
        print(f"Horario cargado manualmente desde '{self.archivo_excel}' con éxito.")

    def _crear_backup(self):
        self.backup_horarios = {
            "grupos": copy.deepcopy(self.horarios_grupos),
            "maestros": copy.deepcopy(self.horarios_maestros),
        }
        print("Backup del estado actual creado.")

    def _restaurar_backup(self):
        if self.backup_horarios:
            self.horarios_grupos = self.backup_horarios["grupos"]
            self.horarios_maestros = self.backup_horarios["maestros"]
            print(
                "\n¡Operación cancelada! El horario ha sido restaurado a su estado original."
            )
            self.backup_horarios = None

    def _get_clase_info(self, grupo, dia, periodo):
        df_grupo = self.horarios_grupos.get(grupo)
        if df_grupo is None:
            print(
                f"Advertencia: No se encontró el grupo '{grupo}' en los datos cargados."
            )
            return None

        clase_str = df_grupo.at[periodo, dia]
        if clase_str == VALOR_VACIO:
            return None

        partes = str(clase_str).split("\n")
        materia = partes[0]
        aula = partes[1] if len(partes) > 1 else ""

        maestro = None
        for maestro_nombre, df_maestro in self.horarios_maestros.items():
            if df_maestro.at[periodo, dia] == grupo:
                maestro = maestro_nombre
                break

        if maestro is None:
            pass

        return {
            "grupo": grupo,
            "materia": materia,
            "aula": aula,
            "maestro": maestro,
            "dia": dia,
            "periodo": periodo,
        }

    def _seleccionar_maestro(self):
        maestros = sorted(list(self.horarios_maestros.keys()))
        while True:
            limpiar_pantalla()
            print("\n--- Seleccione un Maestro ---")
            for i, maestro in enumerate(maestros):
                print(f"[{i+1}] {maestro}")
            try:
                eleccion = int(input("Ingrese el número del maestro: ")) - 1
                if 0 <= eleccion < len(maestros):
                    return maestros[eleccion]
                print("Número fuera de rango.")
            except (ValueError, IndexError):
                print("Entrada no válida.")

    def _seleccionar_clase_origen(self, maestro_nombre):
        df_maestro = self.horarios_maestros[maestro_nombre]
        while True:
            limpiar_pantalla()
            print(f"--- Horario Actual de {maestro_nombre} ---")
            self._mostrar_horario(df_maestro)
            print("\n" + "=" * 30)
            print("Seleccione la CLASE que desea mover.")
            try:
                dia_idx = (
                    int(input(f"Ingrese el número del día (1-Lunes, ..., 5-Viernes): "))
                    - 1
                )
                dia_nombre = DIAS[dia_idx]
                periodo_idx = (
                    int(
                        input(
                            f"Ingrese el número del período (1 a {PERIODOS_POR_DIA}): "
                        )
                    )
                    - 1
                )
                if not (
                    0 <= dia_idx < len(DIAS) and 0 <= periodo_idx < PERIODOS_POR_DIA
                ):
                    input("Valores fuera de rango. Presione Enter para reintentar...")
                    continue
                grupo_nombre = df_maestro.at[periodo_idx, dia_nombre]
                if grupo_nombre == VALOR_VACIO or grupo_nombre is None:
                    input(
                        "\nHa seleccionado un espacio vacío. Debe seleccionar una clase existente. Presione Enter..."
                    )
                    continue

                clase_info = self._get_clase_info(grupo_nombre, dia_nombre, periodo_idx)
                if clase_info is None:
                    input(
                        "\nError al obtener información de la clase. Presione Enter..."
                    )
                    continue

                clase_info["dia_origen"] = dia_nombre
                clase_info["periodo_origen"] = periodo_idx
                return clase_info
            except (ValueError, IndexError, KeyError):
                input(
                    "Entrada no válida o error de datos. Presione Enter para reintentar..."
                )

    def _seleccionar_slot_destino(self):
        while True:
            print("\n--- Seleccione el Día y Período de Destino ---")
            print("(Ingrese 'c' en cualquier momento para cancelar la selección)")
            try:
                dia_input = input(
                    f"Ingrese el número del día de destino (1-Lunes, ..., 5-Viernes): "
                )
                if dia_input.lower() == "c":
                    return None, None

                dia_idx = int(dia_input) - 1
                if not (0 <= dia_idx < len(DIAS)):
                    print("Número de día fuera de rango.")
                    continue

                periodo_input = input(
                    f"Ingrese el número del período de destino (1 a {PERIODOS_POR_DIA}): "
                )
                if periodo_input.lower() == "c":
                    return None, None

                periodo_idx = int(periodo_input) - 1
                if not (0 <= periodo_idx < PERIODOS_POR_DIA):
                    print("Número de período fuera de rango.")
                    continue
                return DIAS[dia_idx], periodo_idx
            except (ValueError, IndexError):
                print("Entrada no válida.")

    def _mover_clase_recursivo(
        self, clase_info, dia_dest, periodo_dest, grupo, nivel=0
    ):
        if nivel > 15:
            print("¡Error! Demasiados desplazamientos recursivos. Operación cancelada.")
            return None

        maestro = clase_info["maestro"]
        dia_orig = clase_info["dia_origen"]
        periodo_orig = clase_info["periodo_origen"]
        materia = clase_info["materia"]
        aula = clase_info["aula"]

        clase_dest_info = self._get_clase_info(grupo, dia_dest, periodo_dest)

        if clase_dest_info is None:
            movimiento = self._mover_clase_simple(
                grupo,
                maestro,
                dia_orig,
                periodo_orig,
                dia_dest,
                periodo_dest,
                materia,
                aula,
            )
            return [movimiento]

        else:
            limpiar_pantalla()
            print("\n¡CONFLICTO DE HORARIO!")
            print(
                f"Intentando mover '{materia}' de {maestro} a [{dia_dest}, P{periodo_dest+1}]"
            )
            print(f"Pero el slot ya está ocupado por:")
            print(f"  - Materia: {clase_dest_info['materia']}")
            print(f"  - Grupo:   {clase_dest_info['grupo']}")
            print(f"  - Maestro: {clase_dest_info['maestro']}")

            print("\n¿Qué desea hacer?")
            print(
                "[1] Mover la clase conflictiva a otro lugar para liberar el espacio."
            )
            print(
                "[2] Cancelar TODA la operación (deshacer todos los cambios realizados)."
            )

            while True:
                opcion = input("Seleccione una opción: ")
                if opcion == "1":
                    break
                elif opcion == "2":
                    return None
                else:
                    print("Opción no válida.")

            print("\nDebe seleccionar un NUEVO DESTINO para la clase desplazada.")
            input("Presione Enter para seleccionar el nuevo destino...")

            nuevo_dia, nuevo_periodo = self._seleccionar_slot_destino()

            if nuevo_dia is None:
                return None

            clase_dest_info["dia_origen"] = dia_dest
            clase_dest_info["periodo_origen"] = periodo_dest

            movimientos_recursivos = self._mover_clase_recursivo(
                clase_dest_info, nuevo_dia, nuevo_periodo, grupo, nivel + 1
            )

            if movimientos_recursivos is None:
                return None

            movimiento_actual = self._mover_clase_simple(
                grupo,
                maestro,
                dia_orig,
                periodo_orig,
                dia_dest,
                periodo_dest,
                materia,
                aula,
            )
            return [movimiento_actual] + movimientos_recursivos

    def _mover_clase_simple(
        self,
        grupo,
        maestro,
        dia_orig,
        periodo_orig,
        dia_dest,
        periodo_dest,
        materia,
        aula,
    ):
        """Mueve una clase y devuelve un registro del movimiento."""
        df_grupo = self.horarios_grupos[grupo]
        df_maestro = self.horarios_maestros[maestro]

        clase_valor = f"{materia}\n{aula}" if aula else materia

        df_grupo.at[periodo_orig, dia_orig] = VALOR_VACIO
        df_maestro.at[periodo_orig, dia_orig] = VALOR_VACIO
        df_grupo.at[periodo_dest, dia_dest] = clase_valor
        df_maestro.at[periodo_dest, dia_dest] = grupo

        return {
            "maestro": maestro,
            "grupo": grupo,
            "materia": materia,
            "dia_origen": dia_orig,
            "periodo_origen": periodo_orig,
            "dia_destino": dia_dest,
            "periodo_destino": periodo_dest,
        }

    def mover_clase_cascada(self):
        self._crear_backup()

        try:
            maestro_seleccionado = self._seleccionar_maestro()
            if not maestro_seleccionado:
                self._restaurar_backup()
                return

            clase_origen = self._seleccionar_clase_origen(maestro_seleccionado)
            if not clase_origen:
                self._restaurar_backup()
                return

            limpiar_pantalla()
            print("--- CLASE A MOVER (VERIFICACIÓN) ---")
            print(f"Grupo: {clase_origen['grupo']}")
            print(f"Materia: {clase_origen['materia']}")
            print(f"Aula: {clase_origen['aula']}")
            print(f"Maestro: {clase_origen['maestro']}")
            print(
                f"Posición: [{clase_origen['dia_origen']}, P{clase_origen['periodo_origen'] + 1}]"
            )
            print(f"\n--- Horario del Grupo {clase_origen['grupo']} ---")
            self._mostrar_horario(self.horarios_grupos[clase_origen["grupo"]])
            input("\nPresione Enter para continuar y seleccionar el destino...")

            grupo = clase_origen["grupo"]
            d2, p2 = self._seleccionar_slot_destino()

            if d2 is None:
                self._restaurar_backup()
                input("\nPresione Enter para volver al menú principal...")
                return

            d1, p1 = clase_origen["dia_origen"], clase_origen["periodo_origen"]

            if (d1, p1) == (d2, p2):
                input(
                    "\nHa seleccionado el mismo slot. No se realiza ninguna acción. Presione Enter..."
                )
                self._restaurar_backup()
                return

            movimientos_realizados = self._mover_clase_recursivo(
                clase_origen, d2, p2, grupo
            )

            limpiar_pantalla()
            if movimientos_realizados is None:
                self._restaurar_backup()
            else:
                self.backup_horarios = None
                print("¡OPERACIÓN REALIZADA CON ÉXITO!\n")
                print("=" * 15, "RESUMEN DE CAMBIOS", "=" * 15)

                maestros_afectados = set()
                for i, mov in enumerate(movimientos_realizados):
                    clase_str = f"{mov['materia']} ({mov['grupo']})"
                    origen_str = f"[{mov['dia_origen']}, P{mov['periodo_origen']+1}]"
                    destino_str = f"[{mov['dia_destino']}, P{mov['periodo_destino']+1}]"
                    print(
                        f"{i+1}. Se movió '{clase_str}' del Maestro '{mov['maestro']}'"
                    )
                    print(f"   Desde: {origen_str} -> Hacia: {destino_str}")
                    maestros_afectados.add(mov["maestro"])

                print("\n" + "=" * 12, "HORARIOS ACTUALIZADOS", "=" * 12)
                for maestro in sorted(list(maestros_afectados)):
                    print(f"\n--- Nuevo Horario de {maestro} ---")
                    self._mostrar_horario(self.horarios_maestros[maestro])

        except Exception as e:
            print(f"\nOcurrió un error inesperado: {e}")
            self._restaurar_backup()

        input("\nPresione Enter para volver al menú principal...")

    def _mostrar_horario(self, df):
        if df.empty:
            print("Horario no disponible o vacío.")
            return
        df_display = df.reindex(columns=COLUMNAS_HORARIO, fill_value=VALOR_VACIO)
        print(df_display.to_string(index=False))

    def guardar_en_excel(self):
        nombre_base = input(
            "\nIngrese el nombre para el nuevo archivo Excel (ej. horarios_modificados.xlsx): "
        )
        if not nombre_base:
            print("Nombre de archivo no válido. Operación cancelada.")
            input("Presione Enter para continuar...")
            return

        if not nombre_base.endswith(".xlsx"):
            nombre_base += ".xlsx"
        ruta_absoluta = os.path.abspath(nombre_base)
        print(f"Guardando horarios en: {ruta_absoluta}")
        try:
            with pd.ExcelWriter(ruta_absoluta, engine="openpyxl") as writer:
                for grupo, df in self.horarios_grupos.items():
                    df[COLUMNAS_HORARIO].to_excel(
                        writer, sheet_name=f"Grupo {grupo}", index=False
                    )
                for maestro, df in self.horarios_maestros.items():
                    df[COLUMNAS_HORARIO].to_excel(
                        writer, sheet_name=sanitizar_nombre_hoja(maestro), index=False
                    )
                if self.df_analisis is not None:
                    self.df_analisis.to_excel(
                        writer, sheet_name="Analisis Horas Servicio", index=False
                    )
            print(
                f"\n¡Archivo guardado con éxito!\nLo puedes encontrar en: {ruta_absoluta}"
            )
        except PermissionError:
            print(
                f"\nERROR: Permiso denegado. No se pudo guardar el archivo en '{ruta_absoluta}'."
            )
            print("Asegúrate de que el archivo no esté abierto en otro programa.")
        except Exception as e:
            print(f"\nOcurrió un error inesperado al guardar el archivo: {e}")
        input("Presione Enter para continuar...")

    def run(self):
        while True:
            limpiar_pantalla()
            print("===== Editor de Horarios Escolares =====")
            print("[1] Mover / Intercambiar una clase")
            print("[2] Guardar cambios en un nuevo archivo")
            print("[3] Salir")
            opcion = input("Seleccione una opción: ")
            if opcion == "1":
                self.mover_clase_cascada()
            elif opcion == "2":
                self.guardar_en_excel()
            elif opcion == "3":
                print("Saliendo del programa.")
                break
            else:
                input("Opción no válida. Presione Enter para intentar de nuevo.")


if __name__ == "__main__":
    archivo_a_editar = "horario_final.xlsx"
    if not os.path.exists(archivo_a_editar):
        print(
            f"Error: El archivo '{archivo_a_editar}' no se encuentra en el directorio actual."
        )
        print(
            "Por favor, asegúrate de que el archivo exista y vuelve a ejecutar el script."
        )
        exit()
    editor = EditorHorarios(archivo_a_editar)
    editor.run()
